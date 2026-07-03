from odoo import fields, _
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO
from odoo import http
from odoo.http import request, content_disposition
from odoo.tools import html2plaintext
from datetime import datetime
from odoo.exceptions import ValidationError
from openpyxl.drawing.image import Image
import base64


class ListOfActiveCasesReport(http.Controller):

    """ List of all active cases """
    @http.route('/list_of_active_cases/download/xlsx', type='http', auth='user')
    def download_list_of_active_cases_report(self, wizard_id=None):
        if not request.env.company.fy_start_date:
            raise ValidationError('Please configure company financial year start date.')

        x_wizard = request.env['cms.reports.wizards'].sudo()
        wizard = x_wizard.browse(int(wizard_id)) if wizard_id else x_wizard
        investigation_obj = request.env['project.project'].sudo()
        peer_review_obj = request.env['forensic.peer.review'].sudo()
        date_today = fields.Datetime.now()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Active Case Report"
        wizard._get_common_headers(sheet, 'B1', 'J13', 'K1', 'M13')

        wizard.create_report_title(sheet, 'B16', 'M17', 'LIST OF ALL ACTIVE CASES')

        columns = [
            ('A', '', 3, ''),
            ('B', 'Sl No.', 8, 'center'),
            ('C', 'Case ID', 22, 'center'),
            ('D', 'Authorization Date', 20, 'center'),
            ('E', 'Case Age (Days)', 18, 'center'),
            ('F', 'Nature of Allegation', 25, 'left'),
            ('G', 'Case Description', 50, 'left'),
            ('H', 'Investigation Manager', 25, 'left'),
            ('I', 'PFO', 25, 'left'),
            ('J', 'Lead Investigator', 20, 'left'),
            ('K', 'Current Status', 20, 'center'),
            ('L', 'Finalization Quarter', 22, 'center'),
            ('M', 'Finalization Month', 22, 'center'),
        ]

        header_row = 19
        wizard.create_table_header(sheet, columns, header_row=header_row)

        row = header_row + 1
        count = 1

        pl_stage_id = request.env.ref('project.project_project_stage_0').id
        cl_stage_id = request.env.ref('project.project_project_stage_3').id
        pfo_job_id = request.env.ref('in2it_forensic_services.principle_of_forensic_officer')

        records = investigation_obj.sudo().search([('case_id', '!=', False), ('stage_id', 'not in', [pl_stage_id, cl_stage_id])])

        for rec in records:
            quarter_label = False
            if rec.final_report_upload_date:
                quarter_label = wizard.get_fy_quarter(rec.final_report_upload_date)

            sheet[f'B{row}'] = count
            sheet[f'C{row}'] = rec.name or ''
            sheet[f'D{row}'] = rec.create_date.strftime('%d-%b-%Y') or ''
            sheet[f'E{row}'] = (date_today - rec.create_date).days or ''
            sheet[f'F{row}'] = rec.allegation_nature_id.name if rec.allegation_nature_id else ''
            sheet[f'G{row}'] = html2plaintext(rec.description or '')
            sheet[f'H{row}'] = rec.user_id.name
            sheet[f'I{row}'] = ', '.join(rec.assigned_forensic_team_id.forensic_member_ids.filtered(
                                lambda m: m.job_id == pfo_job_id).mapped('member_ids.name'))\
                                if rec.assignment_type == 'internal' else rec.assigned_pfo_id.name
            sheet[f'J{row}'] = ', '.join(rec.assigned_forensic_team_id.project_investigator_ids.mapped('name'))\
                                if rec.assignment_type == 'internal' else ''
            sheet[f'K{row}'] = rec.stage_id.name
            sheet[f'L{row}'] = quarter_label if quarter_label else ''
            sheet[f'M{row}'] = rec.final_report_upload_date.strftime("%B %Y") if rec.final_report_upload_date else ''

            wizard.apply_row_styles(sheet, columns, row=row)

            row += 1
            count += 1

        stream = BytesIO()
        workbook.save(stream)
        file_data = stream.getvalue()

        return request.make_response(
            file_data,
            headers=[
                ('Content-Type',
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition('active_case_reports.xlsx'))
            ]
        )

    """ List of contracted cases """
    @http.route('/list_of_contracted_cases/download/xlsx', type='http', auth='user')
    def download_list_of_contracted_cases_report(self, wizard_id=None):
        x_wizard = request.env['cms.reports.wizards'].sudo()
        wizard = x_wizard.browse(int(wizard_id)) if wizard_id else x_wizard
        investigation_obj = request.env['project.project'].sudo()
        currency = request.env.company.currency_id

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Contracted Case Register"
        wizard._get_common_headers(sheet, 'B1', 'N13', 'O1', 'Q13')

        wizard.create_report_title(sheet, 'B16', 'Q17', 'LIST OF ALL CONTRACTED CASES')

        columns = [
            ('A', '', 3, ''),
            ('B', 'Sl No.', 8, 'center'),
            ('C', 'Case ID', 22, 'center'),
            ('D', 'Case Title', 40, 'left'),
            ('E', 'Case Type', 12, 'center'),
            ('F', 'Directorate', 35, 'left'),
            ('G', 'Authorization Date', 20, 'center'),
            ('H', 'Current Case Status', 30, 'center'),
            ('I', 'Vendor Name', 30, 'center'),
            ('J', 'Vendor Contact Mobile', 25, 'center'),
            ('K', 'Vendor Contact email', 30, 'center'),
            ('L', 'SLA Terms', 20, 'center'),
            ('M', 'Assignment Date', 20, 'center'),
            ('N', 'Contract Start Date', 20, 'center'),
            ('O', 'Contract End Date', 20, 'center'),
            ('P', 'Assigned PFO', 25, 'center'),
            ('Q', 'Approved Budget', 20, 'center'),
        ]

        header_row = 19
        wizard.create_table_header(sheet, columns, header_row=header_row)

        row = header_row + 1
        count = 1

        records = investigation_obj.search([('case_id', '!=', False), ('assignment_type', '=', 'external')])
        vendor_line_obj = request.env['investigation.vendor.line'].sudo()
        vendor_tor_obj = request.env['investigation.vendor.tor'].sudo()

        for rec in records:
            outsource_id = vendor_line_obj.search([('project_id', '=', rec.id),('status', '=', 'accept')], limit=1)
            vendor_tor_id = vendor_tor_obj.search([('project_id', '=', rec.id),('status', '=', 'approved')], limit=1)
            start_date = outsource_id.action_date.strftime('%d-%b-%Y') if outsource_id and outsource_id.action_date else ''

            sheet[f'B{row}'] = count
            sheet[f'C{row}'] = rec.name or ''
            sheet[f'D{row}'] = rec.case_id.name or ''
            sheet[f'E{row}'] = rec.assignment_type_id.name or ''
            sheet[f'F{row}'] = rec.directorate_id.name or ''
            sheet[f'G{row}'] = rec.create_date.strftime('%d-%b-%Y') if rec.create_date else ''
            sheet[f'H{row}'] = rec.stage_id.name or ''
            sheet[f'I{row}'] = outsource_id.partner_id.name if outsource_id and outsource_id.partner_id else ''
            sheet[f'J{row}'] = outsource_id.mobile if outsource_id and outsource_id.mobile else ''
            sheet[f'K{row}'] = outsource_id.email if outsource_id and outsource_id.email else ''
            sheet[f'L{row}'] = 'Accepted'
            sheet[f'M{row}'] = start_date
            sheet[f'N{row}'] = start_date
            sheet[f'O{row}'] = rec.final_report_upload_date.strftime('%d-%b-%Y') if rec.final_report_upload_date else ''
            sheet[f'P{row}'] = rec.assigned_pfo_id.name or ''

            budger_cell = sheet[f'Q{row}']
            budger_cell.value = vendor_tor_id.total_amount if vendor_tor_id else 0.00

            symbol = currency.symbol or ''
            budger_cell.number_format = f'"{symbol} "#,##,##0.00'

            wizard.apply_row_styles(sheet, columns, row=row)

            row += 1
            count += 1

        stream = BytesIO()
        workbook.save(stream)
        file_data = stream.getvalue()

        return request.make_response(
            file_data,
            headers=[
                ('Content-Type',
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition('contracted_cases.xlsx'))
            ]
        )

    """ List of control issues report """
    @http.route('/list_of_control_issues/download/xlsx', type='http', auth='user')
    def download_list_of_control_issues_report(self, wizard_id=None):
        x_wizard = request.env['cms.reports.wizards'].sudo()
        wizard = x_wizard.browse(int(wizard_id)) if wizard_id else x_wizard
        investigation_obj = request.env['project.project'].sudo()
        styles = wizard._get_excel_styles()
        fy_year = wizard.financial_year
        company = request.env.user.company_id
        fy_start_month = company.fy_start_date.month
        fy_start_day = company.fy_start_date.day
        quarters = wizard.get_fiscal_quarters(fy_year, fy_start_month, fy_start_day)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Control Issues Report"
        wizard._get_common_headers(sheet, 'B1', 'E13', 'F1', 'H13')

        wizard.create_report_title(sheet, 'B16', 'H17', 'CONTROL ISSUES REPORT')

        columns = [
            ('A', '', 3, ''),
            ('B', 'Sl no.', 8, 'center'),
            ('C', 'Investigation ID', 20, 'center'),
            ('D', 'Case Title', 40, 'left'),
            ('E', 'Directorate', 35, 'left'),
            ('F', 'Lead Investigator', 25, 'center'),
            ('G', 'Report Date', 20, 'center'),
            ('H', 'Control Issue Description', 70, 'left'),
        ]

        header_row = 19

        wizard.create_table_header(sheet, columns, header_row=header_row)

        row = header_row + 1
        uimp_stage_id = request.env.ref('in2it_project_management.project_project_stage_2_1').id
        if quarters:
            for q_name, (start_date, end_date) in quarters.items():
                sheet.merge_cells(f'B{row + 1}:H{row + 1}')
                # Quarter Header
                cell = sheet[f'B{row + 1}']
                cell.border = styles['border']
                cell.alignment = styles['left_align']
                cell.font = styles['sub_header_font']
                cell.value = f"{q_name} ({start_date.strftime('%b')} - {end_date.strftime('%b')}) {fy_year}"
                row += 3

                records = investigation_obj.search([('case_id', '!=', False),
                                                    ('assignment_type', '=', 'internal'),
                                                    ('stage_id', '=', uimp_stage_id),
                                                    ('report_signoff_date', '>=' , start_date.date()),
                                                    ('report_signoff_date', '<=' , end_date.date())])

                count = 1
                for rec in records:
                    sheet[f'B{row}'] = count
                    sheet[f'C{row}'] = rec.name or ''
                    sheet[f'D{row}'] = rec.case_id.name or ''
                    sheet[f'E{row}'] = rec.directorate_id.name or ''
                    sheet[f'F{row}'] = ', '.join(rec.assigned_forensic_team_id.project_investigator_ids.mapped('name')) or ''
                    sheet[f'G{row}'] = rec.report_signoff_date.strftime('%d-%b-%y') if rec.report_signoff_date else ''
                    sheet[f'H{row}'] = html2plaintext(rec.control_issues or '')

                    wizard.apply_row_styles(sheet, columns, row=row)

                    row += 1
                    count += 1

        stream = BytesIO()
        workbook.save(stream)
        file_data = stream.getvalue()

        return request.make_response(
            file_data,
            headers=[
                ('Content-Type',
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition('control_issues_reports.xlsx'))
            ]
        )

    """ List of evidence report register"""
    @http.route('/list_of_evidence_report_register/download/xlsx', type='http', auth='user')
    def download_list_of_evidence_report_register(self, wizard_id=None):
        x_wizard = request.env['cms.reports.wizards'].sudo()
        wizard = x_wizard.browse(int(wizard_id)) if wizard_id else x_wizard
        evidence_obj = request.env['complaint.physical.item'].sudo()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Evidence Report Register"
        wizard._get_common_headers(sheet, 'B1', 'K13', 'L1', 'N13')

        wizard.create_report_title(sheet, 'B16', 'N17', 'EVIDENCE REPORT REGISTER')

        columns = [
            ('A', '', 3, ''),
            ('B', 'Sl no.', 8, 'center'),
            ('C', 'Investigation ID', 20, 'center'),
            ('D', 'Case Title', 35, 'left'),
            ('E', 'Item Title', 35, 'left'),
            ('F', 'Item Category', 25, 'center'),
            ('G', 'Quantity', 15, 'center'),
            ('H', 'Unit of Measures', 20, 'center'),
            ('I', 'Recovery Location', 25, 'left'),
            ('J', 'Storage Location', 25, 'left'),
            ('K', 'Status', 20, 'center'),
            ('L', 'Document Type', 20, 'center'),
            ('M', 'Attachment Name', 25, 'left'),
            ('N', 'Is Exhibit', 15, 'center'),
        ]

        wizard.create_table_header(sheet, columns, header_row=19)

        row = 20
        count = 1

        planning_stage = request.env.ref('project.project_project_stage_0').id
        evidence_ids = evidence_obj.search([('project_id', '!=', False), ('project_id.stage_id', '!=', planning_stage)])
        for rec in evidence_ids:
            sheet[f'B{row}'] = count
            sheet[f'C{row}'] = rec.project_id.name or ''
            sheet[f'D{row}'] = rec.project_id.case_id.name or ''
            sheet[f'E{row}'] = rec.item_title or ''
            sheet[f'F{row}'] = rec.item_category or ''
            sheet[f'G{row}'] = rec.quantity or ''
            sheet[f'H{row}'] = rec.uom_id.name or ''
            sheet[f'I{row}'] = rec.recovery_location or ''
            sheet[f'J{row}'] = rec.storage_location or ''
            sheet[f'K{row}'] = rec.status or ''
            sheet[f'L{row}'] = rec.doc_type_id.name or ''
            sheet[f'M{row}'] = ', '.join(rec.attachment_ids.mapped('name')) if rec.attachment_ids else ''
            sheet[f'N{row}'] = 'Yes' if rec.is_exhibit else 'No'

            wizard.apply_row_styles(sheet, columns, row=row)

            row += 1
            count += 1

        stream = BytesIO()
        workbook.save(stream)
        file_data = stream.getvalue()

        return request.make_response(
            file_data,
            headers=[
                ('Content-Type',
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition('evidence_report_register.xlsx'))
            ]
        )

    """ SAPS / BEE Register Report """
    @http.route('/saps_bee_register_report/download/xlsx', type='http', auth='user')
    def download_saps_bee_register_report(self, wizard_id=None):
        x_wizard = request.env['cms.reports.wizards'].sudo()
        wizard = x_wizard.browse(int(wizard_id)) if wizard_id else x_wizard
        recom_obj = request.env['project.category'].sudo()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "SAPS-BEE Register Report"
        wizard._get_common_headers(sheet, 'B1', 'E13', 'F1', 'H13')

        wizard.create_report_title(sheet, 'B16', 'H17', 'SAPS-BEE REGISTER REPORT')

        columns = [
            ('A', '', 3, ''),
            ('B', 'Sl no.', 8, 'center'),
            ('C', 'Investigation ID', 25, 'center'),
            ('D', 'Case Title', 40, 'left'),
            ('E', 'Directorate', 40, 'left'),
            ('F', 'Lead Investigator', 25, 'center'),
            ('G', 'Recommendation Category', 30, 'center'),
            ('H', 'Recommendation Description', 50, 'left'),
        ]

        wizard.create_table_header(sheet, columns, header_row=19)

        row = 20
        count = 1

        saps_categ_id = request.env.ref('in2it_project_management.recom_saps', raise_if_not_found=False)
        bee_categ_id = request.env.ref('in2it_project_management.compcom_bbee', raise_if_not_found=False)
        completed_stage = request.env.ref('project.project_project_stage_3', raise_if_not_found=False)

        records = recom_obj.search([
            ('project_id', '!=', False),
            ('project_id.stage_id', '=', completed_stage.id),
            ('category_id', 'in', [saps_categ_id.id, bee_categ_id.id])
        ])

        saps_count = 0
        bee_count = 0
        for rec in records:
            sheet[f'B{row}'] = count
            sheet[f'C{row}'] = rec.project_id.name or ''
            sheet[f'D{row}'] = rec.project_id.case_id.name or ''
            sheet[f'E{row}'] = rec.project_id.directorate_id.name or ''
            sheet[f'F{row}'] = ', '.join(rec.project_id.assigned_forensic_team_id.project_investigator_ids.mapped('name')) or ''
            sheet[f'G{row}'] = rec.category_id.name or ''
            sheet[f'H{row}'] = rec.category_description or ''

            if rec.category_id == saps_categ_id:
                saps_count += 1
            if rec.category_id == bee_categ_id:
                bee_count += 1

            wizard.apply_row_styles(sheet, columns, row=row)

            row += 1
            count += 1


        cell = sheet[f'C{row + 3}']
        cell.value = 'Category Summary'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        summary_columns = [
            ('A', '', 3, ''),
            ('B', '', 8, ''),
            ('C', 'Recommendation Category', 30, 'center'),
            ('D', 'Count of Recommendations', 40, 'center')
        ]
        wizard.create_table_header(sheet, summary_columns, header_row=row+5)

        row += 6
        sheet[f'C{row}'] = 'SAPS'
        sheet[f'D{row}'] = saps_count

        styles = wizard._get_excel_styles()
        sheet[f'C{row}'].alignment = styles['center_align']
        sheet[f'C{row}'].font = styles['bold']
        sheet[f'D{row}'].alignment = styles['center_align']
        row += 1

        sheet[f'C{row}'] = 'BEE'
        sheet[f'D{row}'] = bee_count

        sheet[f'C{row}'].alignment = styles['center_align']
        sheet[f'C{row}'].font = styles['bold']
        sheet[f'D{row}'].alignment = styles['center_align']

        stream = BytesIO()
        workbook.save(stream)
        file_data = stream.getvalue()

        return request.make_response(
            file_data,
            headers=[
                ('Content-Type',
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition('saps_bee_reports.xlsx'))
            ]
        )

    """ Investigation Report Register """
    @http.route('/inv_report_register/download/xlsx', type='http', auth='user')
    def download_inv_report_register(self, wizard_id=None):
        x_wizard = request.env['cms.reports.wizards'].sudo()
        wizard = x_wizard.browse(int(wizard_id)) if wizard_id else x_wizard
        investigation_obj = request.env['project.project'].sudo()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Inv. Report Register"
        wizard._get_common_headers(sheet, 'B1', 'L13', 'M1', 'O13')

        wizard.create_report_title(sheet, 'B16', 'O17', 'Investigation Report Register')

        columns = [
            ('A', '', 3, ''),
            ('B', 'Sl No.', 8, 'center'),
            ('C', 'Inv Ref', 22, 'center'),
            ('D', 'Complaint Reference', 25, 'center'),
            ('E', 'Case Title', 30, 'left'),
            ('F', 'Nature of Allegation', 25, 'left'),
            ('G', 'Report Date', 20, 'center'),
            ('H', 'Directorate for Action', 30, 'left'),
            ('I', 'Recommendations', 70, 'left'),
            ('J', 'Summary of Findings', 70, 'left'),
            ('K', 'Control Issues', 50, 'left'),
            ('L', 'Investigation Manager', 20, 'center'),
            ('M', 'PFO', 20, 'center'),
            ('N', 'Lead Investigator', 20, 'center'),
            ('O', 'Outcome', 70, 'left'),
        ]

        header_row = 19
        wizard.create_table_header(sheet, columns, header_row=header_row)

        row = header_row + 1
        count = 1

        done_stage = request.env.ref('project.project_project_stage_2')
        records = investigation_obj.search([('case_id', '!=', False),
                                            ('assignment_type', '=', 'internal'),
                                            ('stage_id', '=', done_stage.id)])
        pfo_job_id = request.env.ref('in2it_forensic_services.principle_of_forensic_officer')
        for rec in records:

            sheet[f'B{row}'] = count
            sheet[f'C{row}'] = rec.name or ''
            sheet[f'D{row}'] = rec.case_id.internal_coms_ref or ''
            sheet[f'E{row}'] = rec.case_id.name or ''
            sheet[f'F{row}'] = rec.allegation_nature_id.name or ''
            sheet[f'G{row}'] = rec.report_signoff_date.strftime('%d-%b-%y') if rec.report_signoff_date else ''
            sheet[f'H{row}'] = rec.directorate_id.name or ''
            sheet[f'I{row}'] = "\n".join(f"{rec.category_id.name or ''}\n{rec.category_description or ''}\n"
                                         for rec in rec.recom_project_ids)
            sheet[f'J{row}'] = html2plaintext(rec.findings or '')
            sheet[f'K{row}'] = html2plaintext(rec.control_issues or '')
            sheet[f'L{row}'] = rec.user_id.name or ''
            sheet[f'M{row}'] = ', '.join(rec.assigned_forensic_team_id.forensic_member_ids.filtered(
                                lambda m: m.job_id == pfo_job_id).mapped('member_ids.name'))
            sheet[f'N{row}'] = ', '.join(rec.assigned_forensic_team_id.project_investigator_ids.mapped('name'))
            sheet[f'O{row}'] = html2plaintext(rec.conclusions or '')

            wizard.apply_row_styles(sheet, columns, row=row)

            row += 1
            count += 1


        # --- Add Directorate Summary Sheet ---
        dir_sheet = workbook.create_sheet(title="Directorate Summary")
        wizard._get_common_headers(dir_sheet, 'B1', 'D13', 'E1', 'F13')
        wizard.create_report_title(dir_sheet, 'B16', 'F17', 'Investigation Report Register : Directorate')

        dir_columns = [
            ('A', '', 3, ''),
            ('B', 'Sl No.', 8, 'center'),
            ('C', 'Directorate for Action', 30, 'left'),
            ('D', 'Number of Investigations', 30, 'center'),
            ('E', 'Number of Issues', 20, 'center'),
            ('F', 'Number of Recommendations', 35, 'center'),
        ]

        header_row = 19
        wizard.create_table_header(dir_sheet, dir_columns, header_row=header_row)

        # Aggregate counts by directorate
        records = investigation_obj.search([])
        directorate_data = {}
        for rec in records:
            dir_name = rec.directorate_id.name or 'Unknown'
            if dir_name not in directorate_data:
                directorate_data[dir_name] = {'investigations': 0, 'issues': 0, 'recommendations': 0}

            directorate_data[dir_name]['investigations'] += 1
            directorate_data[dir_name]['issues'] += 1 if rec.control_issues else 0
            directorate_data[dir_name]['recommendations'] += len(rec.recom_project_ids)

        row = header_row + 1
        count = 1
        for dir_name, counts in directorate_data.items():
            dir_sheet[f'B{row}'] = count
            dir_sheet[f'C{row}'] = dir_name
            dir_sheet[f'D{row}'] = counts['investigations']
            dir_sheet[f'E{row}'] = counts['issues']
            dir_sheet[f'F{row}'] = counts['recommendations']

            wizard.apply_row_styles(dir_sheet, dir_columns, row=row)

            row += 1
            count += 1

        stream = BytesIO()
        workbook.save(stream)
        file_data = stream.getvalue()

        return request.make_response(
            file_data,
            headers=[
                ('Content-Type',
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition('inv_report_register.xlsx'))
            ]
        )

    """ Inv MFMA Investigation Register """
    @http.route('/mfma_inv_report/download/xlsx', type='http', auth='user')
    def download_mfma_inv_report(self, wizard_id=None):
        x_wizard = request.env['cms.reports.wizards'].sudo()
        wizard = x_wizard.browse(int(wizard_id)) if wizard_id else x_wizard
        investigation_obj = request.env['project.project'].sudo()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "MFMA Investigation Initiated"
        wizard._get_common_headers(sheet, 'B1', 'D13', 'E1', 'F13')

        wizard.create_report_title(sheet, 'B16', 'F17', 'MFMA CIRCULAR 88 : INVESTIGATIONS INITIATED IN PAST QUARTER')

        columns = [
            ('A', '', 3, ''),
            ('B', 'Sl No.', 8, 'center'),
            ('C', 'Internal Reference', 30, 'center'),
            ('D', 'Referral', 25, 'center'),
            ('E', 'Date authorized by\noffice of the city manager', 30, 'center'),
            ('F', 'CM Reference Number', 30, 'center')
        ]

        header_row = 19
        wizard.create_table_header(sheet, columns, header_row=header_row)

        row = header_row + 1
        count = 1

        records = investigation_obj.search([('case_id', '!=', False),
                                            ('assignment_type', '=', 'internal')])
        for rec in records:
            sheet[f'B{row}'] = count
            sheet[f'C{row}'] = rec.name or ''
            sheet[f'D{row}'] = rec.assignment_type_id.name or ''
            sheet[f'E{row}'] = rec.create_date.strftime('%d-%b-%Y') if rec.create_date else ''
            sheet[f'F{row}'] = '---------'

            wizard.apply_row_styles(sheet, columns, row=row)

            row += 1
            count += 1

        cm_group = request.env.ref('in2it_forensic_services.group_fcm_city_manager')
        cm_user_id = request.env['res.users'].search([('groups_id', 'in', cm_group.id)], limit=1)

        con_row = row + 3
        wizard.create_report_title(sheet, f'B{con_row}', f'C{con_row}', 'CONTACT DETAILS')

        # Name
        sheet.merge_cells(f'B{con_row + 1}:C{con_row + 1}')
        sheet[f'B{con_row + 1}'] = 'NAME'

        sheet.merge_cells(f'D{con_row + 1}:E{con_row + 1}')
        sheet[f'D{con_row + 1}'] = cm_user_id.name if cm_user_id else ''

        # Contact number
        sheet.merge_cells(f'B{con_row + 2}:C{con_row + 2}')
        sheet[f'B{con_row + 2}'] = 'CONTACT NUMBERS'

        sheet.merge_cells(f'D{con_row + 2}:E{con_row + 2}')
        sheet[f'D{con_row + 2}'] = " / ".join(
            filter(None, [cm_user_id.partner_id.phone, cm_user_id.partner_id.mobile]))

        # Email address
        sheet.merge_cells(f'B{con_row + 3}:C{con_row + 3}')
        sheet[f'B{con_row + 3}'] = 'E-MAIL ADDRESS'

        sheet.merge_cells(f'D{con_row + 3}:E{con_row + 3}')
        sheet[f'D{con_row + 3}'] = cm_user_id.partner_id.email or ''

        # Directorate
        sheet.merge_cells(f'B{con_row + 4}:C{con_row + 4}')
        sheet[f'B{con_row + 4}'] = 'DIRECTORATE'

        sheet.merge_cells(f'D{con_row + 4}:E{con_row + 4}')
        sheet[f'D{con_row + 4}'] = 'Office of the City Manager'

        # Signature
        sheet.merge_cells(f'B{con_row + 5}:C{con_row + 5}')
        sheet[f'B{con_row + 5}'] = 'SIGNATURE OF AUTHOR'

        sheet.merge_cells(f'D{con_row + 5}:E{con_row + 9}')
        image_data = cm_user_id.sign_signature

        image_bytes = base64.b64decode(image_data)

        image_stream = BytesIO(image_bytes)

        img = Image(image_stream)

        img.width = 350
        img.height = 70

        sheet.add_image(img, f'D{con_row + 8}')

        #--- Add MFMA Circular Report ---

        mfmac_sheet = workbook.create_sheet(title="MFMA Circular Report")
        wizard._get_common_headers(mfmac_sheet, 'B1', 'D13', 'E1', 'E13')
        wizard.create_report_title(mfmac_sheet, 'B16', 'E17', 'MFMA Circular Report')

        mfmac_columns = [
            ('A', '', 3, ''),
            ('B', 'Sl No.', 8, 'center'),
            ('C', 'Reporting Quarter', 30, 'center'),
            ('D', 'Investigations initiated\n in past quarter', 50, 'center'),
            ('E', 'Investigations conducted\n in past quarter', 50, 'center'),
        ]

        header_row = 19
        wizard.create_table_header(mfmac_sheet, mfmac_columns, header_row=header_row)

        quarter_data = wizard.generate_quarterly_report(investigation_obj)
        row = header_row + 1
        count = 1
        for key in sorted(quarter_data.keys()):
            data = quarter_data[key]

            mfmac_sheet[f'B{row}'] = count
            mfmac_sheet[f'C{row}'] = data['label']
            mfmac_sheet[f'D{row}'] = data['initiated']
            mfmac_sheet[f'E{row}'] = data['completed']

            wizard.apply_row_styles(mfmac_sheet, mfmac_columns, row=row)
            row += 1
            count += 1

        stream = BytesIO()
        workbook.save(stream)
        file_data = stream.getvalue()

        return request.make_response(
            file_data,
            headers=[
                ('Content-Type',
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition('mfma_report_register.xlsx'))
            ]
        )

    """ FY Case Register """
    @http.route('/fy_case_register/download/xlsx', type='http', auth='user')
    def download_fy_case_register_report(self, wizard_id=None):
        x_wizard = request.env['cms.reports.wizards'].sudo()
        wizard = x_wizard.browse(int(wizard_id)) if wizard_id else x_wizard
        investigation_obj = request.env['project.project'].sudo()
        fy_year = wizard.financial_year
        date_today = fields.Datetime.now()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "FY Case Register"
        wizard._get_common_headers(sheet, 'B1', 'S13', 'T1', 'V13')
        title = f"Financial Year Case Register : {fy_year if fy_year else ''}"
        wizard.create_report_title(sheet, 'B16', 'V17', title)

        columns = [
            ('A', '', 3, ''),
            ('B', 'Sl No.', 8, 'center'),
            ('C', 'Investigation ID', 25, 'center'),
            ('D', 'Case Title', 30, 'left'),
            ('E', 'Case Type', 20, 'center'),
            ('F', 'Case Source', 30, 'center'),
            ('G', 'Complainant', 30, 'center'),
            ('H', 'Directorate', 30, 'center'),
            ('I', 'Department', 25, 'center'),
            ('J', 'Suspect', 30, 'left'),
            ('K', 'Witness', 30, 'left'),
            ('L', 'Nature of Allegation', 40, 'left'),
            ('M', 'Case Description', 70, 'left'),
            ('N', 'Case Received Date', 20, 'center'),
            ('O', 'Authorization Date', 20, 'center'),
            ('P', 'Case Authorized By', 25, 'center'),
            ('Q', 'Case Assignment', 30, 'center'),
            ('R', 'Investigation Manager', 30, 'center'),
            ('S', 'Lead Investigator', 30, 'center'),
            ('T', 'Case Status', 20, 'center'),
            ('U', 'Case Closed Date', 20, 'center'),
            ('V', 'Case Age', 20, 'center'),
        ]

        header_row = 19
        wizard.create_table_header(sheet, columns, header_row=header_row)

        domain = []
        if fy_year:
            # Get the company financial year start/end months
            company = request.env.user.company_id
            fy_start_month = company.fy_start_date.month
            fy_start_day = company.fy_start_date.day
            fy_end_month = company.fy_end_date.month
            fy_end_day = company.fy_end_date.day

            start_year, end_year = fy_year.split('-')
            start_date = datetime(int(start_year), fy_start_month, fy_start_day)
            end_date = datetime(int(end_year), fy_end_month, fy_end_day)

            domain = [('create_date', '>=', start_date.date()), ('create_date', '<=', end_date.date())]

        records = investigation_obj.search(domain)
        project_close_stage_id = request.env.ref('project.project_project_stage_3')

        row = header_row + 1
        count = 1
        for rec in records:
            sheet[f'B{row}'] = count
            sheet[f'C{row}'] = rec.name or ''
            sheet[f'D{row}'] = rec.case_id.name or ''
            sheet[f'E{row}'] = rec.assignment_type_id.name or ''
            sheet[f'F{row}'] = rec.case_id.allegation_source_id.name or ''
            sheet[f'G{row}'] = rec.partner_id.name or ''
            sheet[f'H{row}'] = rec.directorate_id.name or ''
            sheet[f'I{row}'] = rec.department_id.name or ''
            sheet[f'J{row}'] = ', '.join(rec.suspect_ids.mapped('name')) or ''
            sheet[f'K{row}'] = ', '.join(rec.witness_ids.mapped('name')) or ''
            sheet[f'L{row}'] = rec.allegation_nature_id.name or ''
            sheet[f'M{row}'] = html2plaintext(rec.findings or '')
            sheet[f'N{row}'] = rec.case_id.create_date.strftime('%d-%b-%Y') if rec.case_id else ''
            sheet[f'O{row}'] = rec.create_date.strftime('%d-%b-%Y') or ''
            sheet[f'P{row}'] = rec.create_uid.name or ''
            sheet[f'Q{row}'] = rec.assignment_type or ''
            sheet[f'R{row}'] = rec.user_id.name or ''
            sheet[f'S{row}'] = ', '.join(rec.assigned_forensic_team_id.project_investigator_ids.mapped('name')) or ''
            sheet[f'T{row}'] = 'Closed' if rec.stage_id == project_close_stage_id else 'Active'
            sheet[f'U{row}'] = 'Closed Date'
            sheet[f'V{row}'] = (date_today - rec.create_date).days or ''

            wizard.apply_row_styles(sheet, columns, row=row)
            row += 1
            count += 1

        # --- Financial Report : Metadata---

        fy_sheet1 = workbook.create_sheet(title="FY Case Register Metadata")
        wizard._get_common_headers(fy_sheet1, 'B1', 'C13', 'D1', 'E13')
        title = f"Financial Year Case Register {fy_year if fy_year else ''}: Case Metadata Summary"
        wizard.create_report_title(fy_sheet1, 'B16', 'E17', title)

        fy_sheet1_columns = [
            ('A', '', 3, ''),
            ('B', 'Sl No.', 8, 'center'),
            ('C', 'Period', 50, 'center'),
            ('D', 'Case Type', 20, 'center'),
            ('E', 'Number of Investigations', 30, 'center'),
        ]

        header_row = 19
        wizard.create_table_header(fy_sheet1, fy_sheet1_columns, header_row=header_row)

        domain = []
        case_types = ["DFO", "FOR", "ETH", "LSA", "PRE", "SPK-FOR", "SPK-REC", "LIN", "REC"]
        if fy_year:
            # Get the company financial year start/end months
            company = request.env.user.company_id
            fy_start_month = company.fy_start_date.month
            fy_start_day = company.fy_start_date.day
            fy_end_month = company.fy_end_date.month
            fy_end_day = company.fy_end_date.day

            start_year, end_year = fy_year.split('-')
            start_date = datetime(int(start_year), fy_start_month, fy_start_day)
            end_date = datetime(int(end_year), fy_end_month, fy_end_day)

            quarters = wizard.get_fiscal_quarters(fy_year, fy_start_month, fy_start_day)

            domain = [('create_date', '>=', start_date.date()), ('create_date', '<=', end_date.date())]
            records = investigation_obj.search(domain)
            row = header_row + 1
            count = 1
            if quarters:
                for q, (start, end) in quarters.items():
                    q_name = f"{q} ({start.strftime('%b')}-{end.strftime('%b')})"
                    # Filter records for this quarter
                    records_in_quarter = records.filtered(lambda r: start <= r.create_date <= end)

                    for type in case_types:
                        # Count records of this type
                        type_count = sum(1 for r in records_in_quarter if r.assignment_type_id.name == type)

                        fy_sheet1[f'B{row}'] = count
                        fy_sheet1[f'C{row}'] = q_name
                        fy_sheet1[f'D{row}'] = type
                        fy_sheet1[f'E{row}'] = type_count

                        wizard.apply_row_styles(fy_sheet1, fy_sheet1_columns, row=row)
                        row += 1
                        count += 1
                    row += 1

        stream = BytesIO()
        workbook.save(stream)
        file_data = stream.getvalue()

        return request.make_response(
            file_data,
            headers=[
                ('Content-Type',
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition('fy_case_register.xlsx'))
            ]
        )

    """ Recommendation Implementation - Gov, ED, Stats """
    @http.route('/rec_implementation/download/xlsx', type='http', auth='user')
    def download_rec_implementation_report(self, wizard_id=None):
        x_wizard = request.env['cms.reports.wizards'].sudo()
        wizard = x_wizard.browse(int(wizard_id)) if wizard_id else x_wizard
        investigation_obj = request.env['project.project'].sudo()
        rec_imp_line_obj = request.env['ed.action.line'].sudo()
        rec_obj = request.env['project.category'].sudo()
        styles = wizard._get_excel_styles()
        fy_year = wizard.financial_year
        company = request.env.user.company_id
        fy_start_month = company.fy_start_date.month
        fy_start_day = company.fy_start_date.day
        quarters = wizard.get_fiscal_quarters(fy_year, fy_start_month, fy_start_day)

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # --- Recommendation Implementation: ED---

        sheet.title = "Rec Implementation - GOV"
        wizard._get_common_headers(sheet, 'B1', 'H13', 'I1', 'K13')
        title = f"  Implementation of Recommendations : Governance Recommendations"
        wizard.create_report_title(sheet, 'B16', 'K17', title)

        columns = [
            ('A', '', 3, ''),
            ('B', 'Sl No.', 8, 'center'),
            ('C', 'Investigation ID', 25, 'center'),
            ('D', 'Case Title', 40, 'left'),
            ('E', 'Directorate', 35, 'left'),
            ('F', 'Lead Investigator', 25, 'center'),
            ('G', 'Governance Unit Member', 30, 'left'),
            ('H', 'Report Date', 25, 'center'),
            ('I', "Management Comments /\n Feedback", 70, 'left'),
            ('J', 'Implementation Status', 25, 'center'),
            ('K', 'Attachment', 35, 'center'),
        ]

        wizard.create_table_header(sheet, columns, header_row=19)
        header_row = 19

        row = header_row + 1
        if quarters:
            for q_name, (start_date, end_date) in quarters.items():
                sheet.merge_cells(f'B{row+1}:K{row+1}')
                # Quarter Header
                cell = sheet[f'B{row + 1}']
                cell.border = styles['border']
                cell.alignment = styles['left_align']
                cell.font = styles['sub_header_font']
                cell.value = f"{q_name} ({start_date.strftime('%b')} - {end_date.strftime('%b')}) {fy_year}"
                row += 3

                records = rec_imp_line_obj.search([
                    ('action', '=', 'action'),
                    ('user_id.groups_id', 'in', [request.env.ref('base.group_user').id]),
                    '|',
                    '&',
                    ('gov_recom_id.project_id.report_signoff_date', '>=', start_date.date()),
                    ('gov_recom_id.project_id.report_signoff_date', '<=', end_date.date()),
                    '&',
                    ('recom_id.project_id.report_signoff_date', '>=', start_date.date()),
                    ('recom_id.project_id.report_signoff_date', '<=', end_date.date()),
                ])

                count = 1
                for rec in records:
                    sheet[f'B{row}'] = count
                    sheet[f'C{row}'] = rec.gov_recom_id.project_id.name or rec.recom_id.project_id.name or ''
                    sheet[f'D{row}'] = rec.gov_recom_id.project_id.case_id.name or rec.recom_id.project_id.case_id.name or ''
                    sheet[f'E{row}'] = rec.gov_recom_id.project_id.directorate_id.name or rec.recom_id.project_id.directorate_id.name or ''
                    sheet[f'F{row}'] = ', '.join(rec.gov_recom_id.project_id.assigned_forensic_team_id.project_investigator_ids.mapped('name')) or\
                                       ', '.join(rec.recom_id.project_id.assigned_forensic_team_id.project_investigator_ids.mapped('name')) or ''
                    sheet[f'G{row}'] = rec.action_by.name or ''
                    sheet[f'H{row}'] = rec.gov_recom_id.project_id.report_signoff_date or rec.recom_id.project_id.report_signoff_date
                    sheet[f'I{row}'] = rec.remark or ''
                    sheet[f'J{row}'] = rec.status or ''
                    sheet[f'K{row}'] = ', \n'.join(rec.attachment_ids.mapped('name')) or ''

                    wizard.apply_row_styles(sheet, columns, row=row)
                    row += 1
                    count += 1

        # --- Recommendation Implementation: ED---

        rec_ed_sheet = workbook.create_sheet(title="Rec Implementation - ED")
        wizard._get_common_headers(rec_ed_sheet, 'B1', 'H13', 'I1', 'K13')
        title = f"Implementation of Recommendations : ED Recommendations"
        wizard.create_report_title(rec_ed_sheet, 'B16', 'K17', title)
        rec_ed_columns = [
            ('A', '', 3, ''),
            ('B', 'Sl No.', 8, 'center'),
            ('C', 'Investigation ID', 25, 'center'),
            ('D', 'Case Title', 40, 'left'),
            ('E', 'Directorate', 35, 'left'),
            ('F', 'Lead Investigator', 25, 'center'),
            ('G', 'ED Name', 30, 'center'),
            ('H', 'Report Date', 25, 'center'),
            ('I', 'Comments', 70, 'left'),
            ('J', 'Implementation Status', 25, 'center'),
            ('K', 'Attachment', 35, 'center'),
        ]


        header_row = 19
        wizard.create_table_header(rec_ed_sheet, rec_ed_columns, header_row=header_row)


        row = header_row + 1

        if quarters:
            for q_name, (start_date, end_date) in quarters.items():
                rec_ed_sheet.merge_cells(f'B{row+1}:K{row+1}')
                # Quarter Header
                cell = rec_ed_sheet[f'B{row + 1}']
                cell.border = styles['border']
                cell.alignment = styles['left_align']
                cell.font = styles['sub_header_font']
                cell.value = f"{q_name} ({start_date.strftime('%b')} - {end_date.strftime('%b')}) {fy_year}"
                row += 3

                records = rec_imp_line_obj.search([
                    ('action', '=', 'action'),
                    ('user_id.groups_id', 'not in', [request.env.ref('base.group_user').id]),
                    '|',
                    '&',
                    ('gov_recom_id.project_id.report_signoff_date', '>=', start_date.date()),
                    ('gov_recom_id.project_id.report_signoff_date', '<=', end_date.date()),
                    '&',
                    ('recom_id.project_id.report_signoff_date', '>=', start_date.date()),
                    ('recom_id.project_id.report_signoff_date', '<=', end_date.date()),
                ])
                count = 1
                for rec in records:
                    rec_ed_sheet[f'B{row}'] = count
                    rec_ed_sheet[f'C{row}'] = rec.recom_id.project_id.name or ''
                    rec_ed_sheet[f'D{row}'] = rec.recom_id.project_id.case_id.name or ''
                    rec_ed_sheet[f'E{row}'] = ', '.join(rec.gov_recom_id.project_id.assigned_forensic_team_id.project_investigator_ids.mapped('name')) or ''
                    rec_ed_sheet[f'F{row}'] = rec.recom_id.project_id.directorate_id.name or ''
                    rec_ed_sheet[f'G{row}'] = rec.recom_id.project_id.directorate_id.ed_officer_id.name or ''
                    rec_ed_sheet[f'H{row}'] = rec.recom_id.project_id.report_signoff_date or ''
                    rec_ed_sheet[f'I{row}'] = rec.remark or ''
                    rec_ed_sheet[f'J{row}'] = rec.status
                    rec_ed_sheet[f'K{row}'] = ', \n'.join(rec.attachment_ids.mapped('name')) or ''

                    wizard.apply_row_styles(rec_ed_sheet, rec_ed_columns, row=row)
                    row += 1
                    count += 1

            # --- Recommendation Implementation: Stats---

        rec_st_sheet = workbook.create_sheet(title="Rec Implementation - Stats")
        wizard._get_common_headers(rec_st_sheet, 'B1', 'D13', 'E1', 'F13')
        title = f"Implementation of Recommendations : Directorate Statistical Summary"
        wizard.create_report_title(rec_st_sheet, 'B16', 'F17', title)
        rec_st_columns = [
            ('A', '', 3, ''),
            ('B', 'Sl No.', 8, 'center'),
            ('C', 'Directorate for Action', 35, 'center'),
            ('D', 'Number of Recommendations', 30, 'center'),
            ('E', 'Implemented', 25, 'center'),
            ('F', 'In Progress', 25, 'center'),
        ]

        header_row = 19
        wizard.create_table_header(rec_st_sheet, rec_st_columns, header_row=header_row)

        row = header_row + 1
        if quarters:
            for q_name, (start_date, end_date) in quarters.items():
                rec_st_sheet.merge_cells(f'B{row+1}:F{row+1}')
                # Quarter Header
                cell = rec_st_sheet[f'B{row + 1}']
                cell.border = styles['border']
                cell.alignment = styles['left_align']
                cell.font = styles['sub_header_font']
                cell.value = f"{q_name} ({start_date.strftime('%b')} - {end_date.strftime('%b')}) {fy_year}"
                row += 3

                records = rec_obj.search([
                    ('assignment_date', '>=', start_date.date()),
                    ('assignment_date', '<=', end_date.date()),
                ])

                grouped_data = {}
                for rec in records:
                    directorate = rec.project_id.directorate_id.name or 'Unknown'

                    if directorate not in grouped_data:
                        grouped_data[directorate] = {
                            'total': 0,
                            'closed': 0,
                            'open': 0
                        }

                    grouped_data[directorate]['total'] += 1

                    if rec.status == 'closed':
                        grouped_data[directorate]['closed'] += 1
                    else:
                        grouped_data[directorate]['open'] += 1

                count = 1
                for directorate, vals in grouped_data.items():
                    rec_st_sheet[f'B{row}'] = count
                    rec_st_sheet[f'C{row}'] = directorate
                    rec_st_sheet[f'D{row}'] = vals['total']
                    rec_st_sheet[f'E{row}'] = vals['closed']
                    rec_st_sheet[f'F{row}'] = vals['open']

                    wizard.apply_row_styles(rec_st_sheet, rec_st_columns, row=row)
                    row += 1
                    count += 1

        stream = BytesIO()
        workbook.save(stream)
        file_data = stream.getvalue()

        return request.make_response(
            file_data,
            headers=[
                ('Content-Type',
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition('rec_implementation.xlsx'))
            ]
        )
