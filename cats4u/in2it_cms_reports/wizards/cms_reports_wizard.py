# -*- encoding: utf-8 -*-
##############################################################################
#
#    In2IT Technologies Pvt. Ltd
#    Copyright (C) 2025 (https://www.in2ittech.com)
#
##############################################################################

from odoo import models, fields, _, api
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from io import BytesIO
from odoo import http
from odoo.http import request
import base64
from odoo.exceptions import ValidationError
from datetime import datetime, timedelta
import calendar


class CmsReportsWizard(models.TransientModel):
    _name = 'cms.reports.wizards'
    _description = 'CMS Reports Wizard'

    report_type_id = fields.Many2one('cms.report.config', string='Report Type', required=True, domain=lambda self: self._get_allowed_reports())
    financial_year = fields.Selection(
        selection="_get_financial_years",
        string="Financial Year"
    )
    quarter = fields.Selection([
        ('q1', 'Q-1'),
        ('q2', 'Q-2'),
        ('q3', 'Q-3'),
        ('q4', 'Q-4')], default=False, string='Quarter')
    is_hide_fyq = fields.Boolean('Hide Financial Year', default=False, compute='compute_is_hide_fyq')
    is_hide_fy = fields.Boolean(default=False, compute='compute_is_hide_fyq')

    report_type = fields.Selection([
        ('yearly', 'Annually'),
        ('half_yearly', 'Semi-Annually'),
        ('quarterly', 'Quarterly')
    ], default='yearly', required=True)

    half = fields.Selection([
        ('h1', '1st Half'),
        ('h2', '2nd Half')
    ], string="Half Year")

    start_date = fields.Date(string='Start Date')
    end_date = fields.Date(string='End Date')

    @api.depends('report_type_id')
    def compute_is_hide_fyq(self):
        self.is_hide_fyq = False
        self.is_hide_fy = False
        if self.report_type_id in [
            self.env.ref('in2it_cms_reports.case_register_per_fy'),
            self.env.ref('in2it_cms_reports.rec_implementation'),
            self.env.ref('in2it_cms_reports.list_of_control_issues_reports')]:
            self.is_hide_fy = True

    def _get_financial_years(self):
        """
        Get financial years
        If month is less than company fy start month then last fy is prev year - current year
        If month is greater than company fy start month then last fy is current year - next year
        """
        if not self.env.company.fy_start_date:
            raise ValidationError(_('Please configure Company Financial Year'))

        current_year = fields.Date.today().year

        month = fields.Date.today().month
        if month >= self.env.company.fy_start_date.month:
            current_year += 1
        else:
            current_year = current_year

        years = []
        for year in range(2021, current_year):
            fy = f"{year}-{str(year + 1)}"
            years.append((fy, fy))
        return years

    def _get_allowed_reports(self):
        if self.env.user.has_group('base.group_system') or self.env.user.has_group(
                'in2it_forensic_services.group_fcm_admin'):
            return []
        else:
            return [
                '|',
                ('group_ids', '=', False),
                ('group_ids', 'in', self.env.user.groups_id.ids)
            ]

    def print_xlsx_reports(self):
        if self.report_type_id == self.env.ref('in2it_cms_reports.list_of_active_cases'):
            return {
                'type': 'ir.actions.act_url',
                'url': '/list_of_active_cases/download/xlsx?wizard_id=%s' % self.id,
                'close': True,
            }

        if self.report_type_id == self.env.ref('in2it_cms_reports.list_of_contracted_cases'):
            return {
                'type': 'ir.actions.act_url',
                'url': '/list_of_contracted_cases/download/xlsx?wizard_id=%s' % self.id,
                'close': True,
            }

        if self.report_type_id == self.env.ref('in2it_cms_reports.list_of_control_issues_reports'):
            company = request.env.user.company_id
            if not self.financial_year:
                raise ValidationError(_('Please select financial year.'))

            if not company.fy_start_date:
                raise ValidationError(_('Please configure Company Financial Year'))

            return {
                'type': 'ir.actions.act_url',
                'url': '/list_of_control_issues/download/xlsx?wizard_id=%s' % self.id,
                'close': True,
            }

        if self.report_type_id == self.env.ref('in2it_cms_reports.investigation_report_register'):
            return {
                'type': 'ir.actions.act_url',
                'url': '/inv_report_register/download/xlsx?wizard_id=%s' % self.id,
                'close': True,
            }

        if self.report_type_id == self.env.ref('in2it_cms_reports.mfma_inv_report'):
            return {
                'type': 'ir.actions.act_url',
                'url': '/mfma_inv_report/download/xlsx?wizard_id=%s' % self.id,
                'close': True,
            }

        if self.report_type_id == self.env.ref('in2it_cms_reports.list_of_evidence_report_register'):
            return {
                'type': 'ir.actions.act_url',
                'url': '/list_of_evidence_report_register/download/xlsx?wizard_id=%s' % self.id,
                'close': True,
            }

        if self.report_type_id == self.env.ref('in2it_cms_reports.saps_bee_register_report'):
            return {
                'type': 'ir.actions.act_url',
                'url': '/saps_bee_register_report/download/xlsx?wizard_id=%s' % self.id,
                'close': True,
            }

        if self.report_type_id == self.env.ref('in2it_cms_reports.case_register_per_fy'):
            company = request.env.user.company_id
            if not self.financial_year:
                raise ValidationError(_('Please select financial year.'))

            if not company.fy_start_date:
                raise ValidationError(_('Please configure Company Financial Year'))

            return {
                'type': 'ir.actions.act_url',
                'url': '/fy_case_register/download/xlsx?wizard_id=%s' % self.id,
                'close': True,
            }

        if self.report_type_id == self.env.ref('in2it_cms_reports.rec_implementation'):
            company = request.env.user.company_id
            if not self.financial_year:
                raise ValidationError('Please select financial year.')

            if not company.fy_start_date:
                raise ValidationError(_('Please configure Company Financial Year'))

            return {
                'type': 'ir.actions.act_url',
                'url': '/rec_implementation/download/xlsx?wizard_id=%s' % self.id,
                'close': True,
            }

        else:
            return {'type': 'ir.actions.act_window_close'}

    def border_range(self, sheet, cell_range, border):
        for row in sheet[cell_range]:
            for cell in row:
                cell.border = border

    def _get_common_headers(self, sheet, logo_start, logo_end, title_start, title_end):
        company = request.env.company
        merge_range = f"{logo_start}:{logo_end}"
        sheet.merge_cells(merge_range)

        logo = company.logo
        if logo:
            logo_data = base64.b64decode(logo)
            img = XLImage(BytesIO(logo_data))
            img.width = 480
            img.height = 280
            img.anchor = 'B1'
            sheet.add_image(img)

        merge_title_range = f"{title_start}:{title_end}"
        sheet.merge_cells(merge_title_range)
        cell = sheet[title_start]

        font_confidential = InlineFont(sz=20, b=True, color="FFFF0000")
        font_rest = InlineFont(sz=18, b=True)

        # Create rich text
        rich_text = CellRichText([
            TextBlock(font_confidential, "                  CONFIDENTIAL      \n"),
            TextBlock(font_rest, "     Office of the City Manager       \nEthics and Forensic Services       ")
        ])
        cell.value = rich_text

        # Alignment
        cell.alignment = Alignment(
            horizontal='right',
            vertical='center',
            wrap_text = True
        )

    def generate_quarterly_report(self, investigation_obj):
        records = investigation_obj.search([('case_id', '!=', False),
                                            ('assignment_type', '=', 'internal')])

        quarter_data = {}
        for rec in records:
            # Count initiated by authorization_date
            if rec.create_date:
                label, q, fy = self.get_fy_quarter(rec.create_date)

                key = (fy, q)

                if key not in quarter_data:
                    quarter_data[key] = {
                        'label': label,
                        'initiated': 0,
                        'completed': 0
                    }

                quarter_data[key]['initiated'] += 1

            # Count completed by peer review signoff date
            signoff_date = rec.report_signoff_date

            if signoff_date:
                label, q, fy = self.get_fy_quarter(signoff_date)

                key = (fy, q)

                if key not in quarter_data:
                    quarter_data[key] = {
                        'label': label,
                        'initiated': 0,
                        'completed': 0
                    }

                quarter_data[key]['completed'] += 1

        return quarter_data

    def _get_excel_styles(self):
        thin = Side(border_style="thin", color="000000")
        return {
            'border_all': Border(top=thin, bottom=thin, left=thin, right=thin),
            'border': Border(left=thin, right=thin, top=thin, bottom=thin),
            'header_fill': PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid"),
            'table_header_fill': PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid"),
            'header_font': Font(size=12, bold=True, color="FFFFFF"),
            'sub_header_font': Font(size=12, bold=True, color="000000"),
            'center_align': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'left_align': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'bold':Font(bold=True)
        }

    def get_fy_quarter(self, date):
        """
        Returns financial quarter and month name for a given date and FY start month.

        """
        month = date.month
        fy_start_month = self.env.company.fy_start_date.month
        shifted_month = (month - fy_start_month) % 12
        quarter = (shifted_month // 3) + 1


        # Determine FY label
        if month >= fy_start_month:
            fy_start_year = date.year
            fy_end_year = date.year + 1
        else:
            fy_start_year = date.year - 1
            fy_end_year = date.year

        # Calculate quarter start and end months (1-12)
        q_start_month = ((quarter - 1) * 3 + fy_start_month - 1) % 12 + 1
        q_end_month = ((q_start_month + 2 - 1) % 12) + 1

        start_month_name = calendar.month_abbr[q_start_month]
        end_month_name = calendar.month_abbr[q_end_month]

        fy_label = f"{fy_start_year}-{fy_end_year}"
        label = f"Q{quarter} {fy_label} ({start_month_name}–{end_month_name})"
        label1 = f"Q{quarter} {str(fy_start_year)[-2:]}-{str(fy_end_year)[-2:]}"
        if self.report_type_id == self.env.ref('in2it_cms_reports.list_of_active_cases'):
            return label1
        return label, quarter, fy_start_year

    def create_report_title(self, sheet, start_cell, end_cell, title):
        """
            Merge cells, set borders, title, font, fill, and alignment.
        """
        styles = self._get_excel_styles()
        sheet.merge_cells(f'{start_cell}:{end_cell}')

        self.border_range(sheet, f'{start_cell}:{end_cell}', styles['border_all'])
        title_cell = sheet[start_cell]
        title_cell.value = title
        title_cell.font = styles['header_font']
        title_cell.fill = styles['header_fill']
        title_cell.alignment = styles['center_align']

    def create_table_header(self, sheet, columns, header_row):
        styles = self._get_excel_styles()
        for col, title, _, _, in columns[1:]:  # skip first column if needed
            cell = sheet[f'{col}{header_row}']
            cell.value = title
            cell.font = styles['header_font']
            cell.alignment = styles['center_align']
            cell.border = styles['border']
            cell.fill = styles['table_header_fill']

        # Set column widths
        for col, _, width, _ in columns:
            sheet.column_dimensions[col].width = width
            sheet.row_dimensions[header_row].height = 30

    def apply_row_styles(self, sheet, columns, row):
        styles = self._get_excel_styles()
        for col, _, _, align in columns[1:]:
            cell = sheet[f'{col}{row}']
            cell.border = styles['border']
            if align == 'center':
                cell.alignment = styles['center_align']
            else:
                cell.alignment = styles['left_align']

    def get_fiscal_quarters(self, fy_year: str, fy_start_month: int, fy_start_day: int):
        """
        Returns fiscal quarters with start and end dates.

        Args:
            fy_year (str): e.g., "2025-2026"
            fy_start_month (int): Fiscal year start month
            fy_end_month (int): Fiscal year end month

        Returns:
            dict: {"Q1": (start_date, end_date), ..., "Q4": (start_date, end_date)}
        """
        start_year, end_year = map(int, fy_year.split('-'))

        # Fiscal year start date
        fy_start_date = datetime(start_year, fy_start_month, fy_start_day)

        quarters = {}
        current_start = fy_start_date

        for i in range(1, 5):
            # Calculate quarter end month and year
            quarter_end_month = (current_start.month + 2 - 1) % 12 + 1
            quarter_end_year = current_start.year + ((current_start.month + 2 - 1) // 12)

            # Calculate last day of quarter_end_month
            next_month = (quarter_end_month % 12) + 1
            next_year = quarter_end_year if next_month != 1 else quarter_end_year + 1
            last_day = (datetime(next_year, next_month, 1) - timedelta(days=1)).day

            quarter_end = datetime(quarter_end_year, quarter_end_month, last_day)

            quarters[f"Q{i}"] = (current_start, quarter_end)

            # Next quarter start
            next_start_month = (current_start.month + 3 - 1) % 12 + 1
            next_start_year = current_start.year + ((current_start.month + 3 - 1) // 12)
            current_start = datetime(next_start_year, next_start_month, 1)

        return quarters
